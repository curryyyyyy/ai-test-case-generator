from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet


HEADER_MAP: list[tuple[str, str]] = [
    ("case_id", "用例ID"),
    ("directory", "模块"),
    ("test_point", "功能点"),
    ("case_level", "优先级"),
    ("precondition", "前置条件"),
    ("steps", "测试步骤"),
    ("expected_result", "预期结果"),
]


def _write_header(worksheet: Worksheet) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_index, (_, header_text) in enumerate(HEADER_MAP, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header_text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def _normalize_cell_value(field_name: str, test_case: dict[str, Any]) -> str:
    raw_value = test_case.get(field_name, "")
    if field_name == "steps":
        if isinstance(raw_value, list):
            return "\n".join(str(item) for item in raw_value)
        return str(raw_value)
    return str(raw_value)


def _write_data_rows(worksheet: Worksheet, test_cases: list[dict[str, Any]]) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_index, test_case in enumerate(test_cases, start=2):
        for col_index, (field_name, _) in enumerate(HEADER_MAP, start=1):
            cell = worksheet.cell(
                row=row_index,
                column=col_index,
                value=_normalize_cell_value(field_name, test_case),
            )
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border


def _auto_adjust_layout(worksheet: Worksheet) -> None:
    for col_index in range(1, len(HEADER_MAP) + 1):
        max_length = 0
        column_letter = worksheet.cell(row=1, column=col_index).column_letter
        for row in worksheet.iter_rows(
            min_col=col_index,
            max_col=col_index,
            min_row=1,
            max_row=worksheet.max_row,
        ):
            cell = row[0]
            cell_text = "" if cell.value is None else str(cell.value)
            longest_line = max((len(line) for line in cell_text.splitlines()), default=0)
            if longest_line > max_length:
                max_length = longest_line

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    worksheet.row_dimensions[1].height = 24
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 36


def export_test_cases_to_excel(
    test_cases: list[dict[str, Any]],
    output_file_path: str | Path,
) -> str:
    """将测试用例列表导出为带样式的 Excel 文件并返回生成路径。"""
    output_path = Path(output_file_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试用例"

    _write_header(worksheet)
    _write_data_rows(worksheet, test_cases)
    _auto_adjust_layout(worksheet)

    workbook.save(output_path)
    workbook.close()
    return str(output_path.resolve())
